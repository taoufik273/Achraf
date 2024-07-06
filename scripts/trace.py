import sqlite3
import os
from openpyxl import load_workbook
from openpyxl.worksheet.pagebreak import Break
from tkinter import Tk, filedialog, messagebox
import win32com.client  # pip install pywin32

def get_absolute_path(*relative_path):
    script_dir = os.path.dirname(__file__)
    return os.path.join(script_dir, *relative_path)

# Paths
db_path = get_absolute_path('..', 'data', 'saisie.db')
input_excel_path = get_absolute_path('..', 'INPUT', 'liste.xlsx')
output_excel_path = get_absolute_path('..', 'OUTPUT', 'trace.xlsx')

# Check if database file exists
if not os.path.exists(db_path):
    raise FileNotFoundError(f"Database file not found: {db_path}")

# Delete the existing output file if it exists
if os.path.exists(output_excel_path):
    os.remove(output_excel_path)

# Connect to the database
conn = sqlite3.connect(db_path)
cursor = conn.cursor()

# Load the Excel template
if not os.path.exists(input_excel_path):
    raise FileNotFoundError(f"Excel template file not found: {input_excel_path}")

workbook = load_workbook(input_excel_path)
sheet = workbook.active

# Rename the active sheet to "TRACE"
sheet.title = "TRACE"

# Fetch data from database
cursor.execute("SELECT annee, centre FROM index2")
index2_data = cursor.fetchone()
annee, centre = index2_data[0], index2_data[1]

cursor.execute("SELECT jour, numliste, heure, ordre, massar, nom FROM liste ORDER BY jour, heure, numliste")
liste_data = cursor.fetchall()

# Organize data by jour, numliste, and heure
organized_data = {}
for row in liste_data:
    jour, numliste, heure, ordre, massar, nom = row
    key = (jour, numliste, heure)
    if key not in organized_data:
        organized_data[key] = []
    organized_data[key].append(row)

# Function to copy and paste headers
def copy_headers(start_row):
    for i in range(1, 6):
        # Copier la hauteur de la ligne d'en-tête
        sheet.row_dimensions[start_row + i - 1].height = sheet.row_dimensions[i].height
        for j in range(1, 9):
            source_cell = sheet.cell(row=i, column=j)
            target_cell = sheet.cell(row=start_row + i - 1, column=j)
            target_cell.value = source_cell.value
            if source_cell.has_style:
                target_cell._style = source_cell._style

# Function to copy styles from template rows
def copy_styles(start_row, end_row, template_start=6):
    for i in range(start_row, end_row + 1):
        template_row = template_start + ((i - start_row) % 31)
        # Copier la hauteur de la ligne
        sheet.row_dimensions[i].height = sheet.row_dimensions[template_row].height
        for j in range(1, 9):
            source_cell = sheet.cell(row=template_row, column=j)
            target_cell = sheet.cell(row=i, column=j)
            if source_cell.has_style:
                target_cell._style = source_cell._style

# Function to copy and paste the specific rows (31 and 32)
def copy_specific_rows(start_row):
    for i in range(31, 33):  # Inclure la ligne 32
        source_row = i
        target_row = start_row + (i - 31)
        # Copier la hauteur de la ligne
        sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
        for j in range(1, 9):
            source_cell = sheet.cell(row=source_row, column=j)
            target_cell = sheet.cell(row=target_row, column=j)
            target_cell.value = source_cell.value  # Copier la valeur du contenu
            if source_cell.has_style:
                target_cell._style = source_cell._style

# Function to set row heights
def set_row_heights(start_row, end_row, height):
    for i in range(start_row, end_row + 1):
        sheet.row_dimensions[i].height = height

# Fill the Excel file
row_index = 1
page_break_count = 0

for (jour, numliste, heure), rows in organized_data.items():
    # Calculate number of pages needed for this list
    pages_needed = (len(rows) - 1) // 25 + 1
    
    for page in range(pages_needed):
        # Copy headers
        copy_headers(row_index)
        
        # Copy specific rows (31 and 32)
        copy_specific_rows(row_index + 30)

        # Write specific header data
        sheet.cell(row=row_index, column=1, value=f"لائحة المترشحين الأحرار لاجتياز اختبارات التربية البدنية لموسم {annee}")
        sheet.cell(row=row_index + 1, column=5, value=centre)
        sheet.cell(row=row_index + 2, column=4, value=jour)
        sheet.cell(row=row_index + 3, column=2, value=numliste)
        sheet.cell(row=row_index + 3, column=6, value=heure)

        # Write data starting from row 6 to 32 for each page
        data_start = row_index + 5
        data_end = data_start + 26  # 32 lines - 6 (starting from 6th row)

        for j in range(data_start, data_end + 1):
            row_index_in_data = page * 25 + (j - data_start)
            if row_index_in_data < len(rows):
                _, _, _, ordre, massar, nom = rows[row_index_in_data]
                sheet.cell(row=j, column=1, value=ordre)
                sheet.cell(row=j, column=2, value=massar)
                sheet.cell(row=j, column=3, value=nom)

        # Copy styles and row heights for the data rows
        copy_styles(data_start, data_end)

        # Add page break after every page except the last one
        page_break_count += 1
        if page_break_count < sum([(len(rows) - 1) // 25 + 1 for rows in organized_data.values()]):
            sheet.row_breaks.append(Break(data_end))

        # Move to the next section
        row_index = data_end + 1

# Save the filled Excel file
workbook.save(output_excel_path)

# Close the database connection
conn.close()

# Convert the Excel file to PDF
def convert_excel_to_pdf(excel_path, pdf_path):
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False

    wb = excel.Workbooks.Open(excel_path)
    wb.ExportAsFixedFormat(0, pdf_path)
    wb.Close()
    excel.Quit()

# Open file dialog to save PDF
root = Tk()
root.withdraw()  # Hide the root window
pdf_path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")])

if pdf_path:
    convert_excel_to_pdf(output_excel_path, pdf_path)
    # Display success message
    messagebox.showinfo("ممتاز", f"تم تحميل بطاقة التنقيط بنجاح")
else:
    messagebox.showwarning("خطأ", "حدث خطأ أثناء التحميل")
