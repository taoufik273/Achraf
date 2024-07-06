import sqlite3
import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import Font
from tkinter import Tk
from tkinter import filedialog, messagebox
import shutil

def get_resource_path(relative_path):
    """Obtenir le chemin absolu vers la ressource, fonctionne pour dev et pour PyInstaller"""
    try:
        # PyInstaller crée une variable temporaire _MEIPASS pour stocker le chemin du bundle
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

# Fonction pour copier les styles
def copy_cell_style(source_cell, target_cell):
    target_cell.font = source_cell.font.copy()
    target_cell.border = source_cell.border.copy()
    target_cell.fill = source_cell.fill.copy()
    target_cell.number_format = source_cell.number_format
    target_cell.protection = source_cell.protection.copy()
    target_cell.alignment = source_cell.alignment.copy()

# Chemins des fichiers
db_path = get_resource_path('data/saisie.db')
excel_path = get_resource_path('INPUT/presence.xlsx')
output_path = get_resource_path('OUTPUT/presence_intervenants.xlsx')

# Connexion à la base de données
conn = sqlite3.connect(db_path)
cursor = conn.cursor()

# Charger le fichier Excel
wb = load_workbook(excel_path)

# Remplir la feuille presence
ws_presence = wb['presence']

# Remplir les cellules spécifiques
# A1
cursor.execute("SELECT annee FROM index2 LIMIT 1")
annee = cursor.fetchone()[0]
ws_presence['A1'] = f"لائحة المتدخلين في اختبار مادة التربية البدنية أحرار لموسم {annee}"

# H2
cursor.execute("SELECT centre FROM index2 LIMIT 1")
centre = cursor.fetchone()[0]
ws_presence['H2'] = centre

# Remplir les colonnes A à F à partir de la ligne 6
cursor.execute("SELECT numero, nom, ppr, travail, code, mission FROM intervenant")
rows = cursor.fetchall()
for i, row in enumerate(rows, start=6):
    for col, value in enumerate(row, start=1):
        cell = ws_presence.cell(row=i, column=col)
        cell.value = value
        # Copier le style de la ligne 6
        copy_cell_style(ws_presence.cell(row=6, column=col), cell)

# Récupérer les jours uniques de la table saisie
cursor.execute("SELECT DISTINCT jour FROM saisie ORDER BY jour")
jours = cursor.fetchall()

# Copier le style de la cellule A5 pour les appliquer aux cellules des jours
source_cell = ws_presence['A5']
font_11 = Font(size=11)

# Remplir les colonnes G5, H5, I5, etc.
for j, jour in enumerate(jours, start=7):
    cell = ws_presence.cell(row=5, column=j)
    cell.value = jour[0]
    copy_cell_style(source_cell, cell)
    cell.font = font_11

    # Reproduire la mise en forme de la colonne A pour les lignes de données
    for i in range(6, len(rows) + 6):
        data_cell = ws_presence.cell(row=i, column=j)
        copy_cell_style(ws_presence[f'A{i}'], data_cell)
        data_cell.font = font_11

# Remplir la feuille intervenants
ws_intervenants = wb['intervenants']

# Remplir les cellules spécifiques
# A1
ws_intervenants['A1'] = f"لائحة المتدخلين في اختبار مادة التربية البدنية أحرار لموسم {annee}"

# G2
ws_intervenants['G2'] = centre

# Récupérer le nombre de jours uniques
cursor.execute("SELECT COUNT(DISTINCT jour) FROM saisie")
nombre_jours = cursor.fetchone()[0]

# Remplir les colonnes A à I à partir de la ligne 7
for i, row in enumerate(rows, start=7):
    for col, value in enumerate(row, start=1):
        cell = ws_intervenants.cell(row=i, column=col)
        cell.value = value
        # Copier le style de la ligne 7
        copy_cell_style(ws_intervenants.cell(row=7, column=col), cell)
    
    # Colonne G: Nombre de jours uniques
    cell_g = ws_intervenants.cell(row=i, column=7)
    cell_g.value = nombre_jours
    copy_cell_style(ws_intervenants.cell(row=7, column=7), cell_g)
    
    # Colonne H: Nombre 2 si le code est 1 ou 2
    cell_h = ws_intervenants.cell(row=i, column=8)
    cell_h.value = 2 if row[4] in [1, 2] else 0
    copy_cell_style(ws_intervenants.cell(row=7, column=8), cell_h)
    
    # Colonne I: Somme de G et H
    cell_i = ws_intervenants.cell(row=i, column=9)
    cell_i.value = cell_g.value + cell_h.value
    copy_cell_style(ws_intervenants.cell(row=7, column=9), cell_i)

# Sauvegarder le fichier Excel dans le dossier OUTPUT
wb.save(output_path)

# Ouvrir la boîte de dialogue de fichier pour permettre à l'utilisateur de sauvegarder une copie du fichier
root = Tk()
root.withdraw()  # Cacher la fenêtre principale

desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], initialfile="presence_intervenants.xlsx", initialdir=desktop_path)

if file_path:
    shutil.copy(output_path, file_path)
    messagebox.showinfo("ممتاز", f"تم تحميل  حضور المتدخلين بنجاح")
else:
    messagebox.showwarning("خطأ", "حدث خطأ أثناء التحميل")


