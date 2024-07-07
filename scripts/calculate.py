import os
import sys
import sqlite3
import openpyxl
import datetime
import logging
from openpyxl.utils.exceptions import InvalidFileException

# Configuration du logging
logging.basicConfig(filename='/home/taoufik273/calculate.log', level=logging.DEBUG, 
                    format='%(asctime)s - %(levelname)s - %(message)s')

def get_absolute_path(relative_path):
    base_path = '/home/taoufik273/Achraf'  # Ajustez ce chemin selon votre structure de dossiers
    return os.path.join(base_path, relative_path)

def delete_temp_excel_files():
    try:
        temp_excel_file_path1 = get_absolute_path('OUTPUT/note_temp.xlsx')
        temp_excel_file_path2 = get_absolute_path('OUTPUT/note.xlsx')
        
        if os.path.exists(temp_excel_file_path1):
            os.remove(temp_excel_file_path1)
            logging.debug(f"Deleted {temp_excel_file_path1}")
        if os.path.exists(temp_excel_file_path2):
            os.remove(temp_excel_file_path2)
            logging.debug(f"Deleted {temp_excel_file_path2}")
    except Exception as e:
        logging.error(f"Error deleting temp files: {str(e)}")

def convertir_valeurs(row):
    return [str(cell) if isinstance(cell, (datetime.time, datetime.date, datetime.datetime)) else cell for cell in row]

def remplir_et_convertir_et_importer():
    try:
        logging.debug("Starting remplir_et_convertir_et_importer")
        
        delete_temp_excel_files()

        output_dir = get_absolute_path('OUTPUT')
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            logging.debug(f"Created directory: {output_dir}")

        db_path = get_absolute_path('data/saisie.db')
        logging.debug(f"Connecting to database: {db_path}")
        conn = sqlite3.connect(db_path)
        c = conn.cursor()

        input_excel_path = get_absolute_path('INPUT/note.xlsx')
        logging.debug(f"Opening Excel file: {input_excel_path}")
        try:
            wb = openpyxl.load_workbook(input_excel_path)
        except InvalidFileException:
            logging.error(f"Invalid Excel file: {input_excel_path}")
            raise
        except FileNotFoundError:
            logging.error(f"Excel file not found: {input_excel_path}")
            raise
        
        sheet = wb['saisie']

        c.execute("SELECT * FROM saisie")
        donnees = c.fetchall()

        ligne = 2
        colonnes_map = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'M', 'O']
        for row in donnees:
            for i, colonne in enumerate(colonnes_map):
                sheet[f"{colonne}{ligne}"] = row[i]
            ligne += 1

        formule_l = sheet['L2'].value
        formule_n = sheet['N2'].value
        formule_p = sheet['P2'].value
        formule_q = sheet['Q2'].value
        formule_r = sheet['R2'].value
        formule_s = sheet['S2'].value
        formule_t = sheet['T2'].value
        formule_u = sheet['U2'].value
        formule_v = sheet['V2'].value

        for row in range(3, sheet.max_row + 1):
            nouvelle_formule_l = formule_l.replace("I2", f"I{row}")
            for cellule in ['J', 'K', 'M']:
                nouvelle_formule_l = nouvelle_formule_l.replace(f"{cellule}2", f"{cellule}{row}")
            sheet[f"L{row}"] = nouvelle_formule_l

            nouvelle_formule_n = formule_n.replace("J2", f"J{row}").replace("M2", f"M{row}").replace("I2", f"I{row}")
            sheet[f"N{row}"] = nouvelle_formule_n

            nouvelle_formule_p = formule_p.replace("J2", f"J{row}").replace("O2", f"O{row}").replace("I2", f"I{row}")
            sheet[f"P{row}"] = nouvelle_formule_p

            nouvelle_formule_q = formule_q.replace("J2", f"J{row}").replace("L2", f"L{row}").replace("N2", f"N{row}").replace("P2", f"P{row}")
            sheet[f"Q{row}"] = nouvelle_formule_q

            nouvelle_formule_r = formule_r.replace("J2", f"J{row}").replace("S2", f"S{row}").replace("L2", f"L{row}").replace("N2", f"N{row}").replace("P2", f"P{row}")
            sheet[f"R{row}"] = nouvelle_formule_r

            nouvelle_formule_s = formule_s.replace("K2", f"K{row}").replace("M2", f"M{row}").replace("O2", f"O{row}").replace("J2", f"J{row}")
            sheet[f"S{row}"] = nouvelle_formule_s

            nouvelle_formule_t = formule_t.replace("R2", f"R{row}")
            sheet[f"T{row}"] = nouvelle_formule_t

            nouvelle_formule_u = formule_u.replace("R2", f"R{row}")
            sheet[f"U{row}"] = nouvelle_formule_u

            nouvelle_formule_v = formule_v.replace("R2", f"R{row}")
            sheet[f"V{row}"] = nouvelle_formule_v

        output_temp_path = get_absolute_path('OUTPUT/note_temp.xlsx')
        logging.debug(f"Saving Excel file: {output_temp_path}")
        wb.save(output_temp_path)

        # Recalculer les formules
        logging.debug("Recalculating formulas")
        wb = openpyxl.load_workbook(output_temp_path, data_only=True)
        sheet = wb['saisie']

        output_final_path = get_absolute_path('OUTPUT/note.xlsx')
        logging.debug(f"Saving recalculated Excel file: {output_final_path}")
        wb.save(output_final_path)

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            row = convertir_valeurs(row)
            c.execute('SELECT COUNT(*) FROM notes WHERE massar = ?', (row[5],))
            if c.fetchone()[0] == 0:
                c.execute('INSERT INTO notes VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)', row)
            else:
                c.execute('''UPDATE notes 
                             SET gym = ?, course = ?, notecourse = ?, poid = ?, notepoid = ?, saut = ?, notesaut = ?, totalnote = ?, moyen = ?, nombrenote = ?, note = ?, absent = ?, dispence = ?
                             WHERE massar = ?''', 
                          (row[9], row[10], row[11], row[12], row[13], row[14], row[15], row[16], row[17], row[18], row[19], row[20], row[21], row[5]))

        conn.commit()
        conn.close()

        logging.debug("Finished remplir_et_convertir_et_importer")
    except Exception as e:
        logging.error(f"Error in remplir_et_convertir_et_importer: {str(e)}", exc_info=True)
        raise

if __name__ == "__main__":
    try:
        remplir_et_convertir_et_importer()
    except Exception as e:
        logging.error(f"Main execution error: {str(e)}", exc_info=True)
        print(f"An error occurred: {str(e)}", file=sys.stderr)
        sys.exit(1)