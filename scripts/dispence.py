import sqlite3
import os
from openpyxl import load_workbook
from openpyxl.worksheet.pagebreak import Break
from openpyxl.styles import Alignment, Font, Border, Side
from tkinter import Tk, filedialog, messagebox
import win32com.client

def get_absolute_path(*relative_path):
    script_dir = os.path.dirname(__file__)
    return os.path.join(script_dir, *relative_path)

# Paths
db_path = get_absolute_path('..', 'data', 'saisie.db')
input_excel_path = get_absolute_path('..', 'INPUT', 'dispence.xlsx')
output_excel_path = get_absolute_path('..', 'OUTPUT', 'dispence_remplie.xlsx')

# Check if database file exists
if not os.path.exists(db_path):
    raise FileNotFoundError(f"Database file not found: {db_path}")

# Delete the existing output file if it exists
if os.path.exists(output_excel_path):
    os.remove(output_excel_path)

# Connect to the database
conn = sqlite3.connect(db_path)
cursor = conn.cursor()

# Load the Excel template
if not os.path.exists(input_excel_path):
    raise FileNotFoundError(f"Excel template file not found: {input_excel_path}")

workbook = load_workbook(input_excel_path)
sheet = workbook.active

# Rename the active sheet to "DSP"
sheet.title = "DSP"

# Fetch data from database
cursor.execute("SELECT annee, centre FROM index2")
index2_data = cursor.fetchone()
annee, centre = index2_data[0], index2_data[1]

cursor.execute("SELECT jour, massar, nom, serie FROM notes WHERE dispence = 'نعم' ORDER BY jour")
dsp_data = cursor.fetchall()

def center_and_bold_cell(cell):
    existing_font = cell.font
    existing_alignment = cell.alignment

    new_font = Font(
        name=existing_font.name,
        size=existing_font.size,
        bold=True,
        italic=existing_font.italic,
        vertAlign=existing_font.vertAlign,
        underline=existing_font.underline,
        strike=existing_font.strike,
        color=existing_font.color
    )

    new_alignment = Alignment(
        horizontal='center',
        vertical='center',
        text_rotation=existing_alignment.text_rotation,
        wrap_text=existing_alignment.wrap_text,
        shrink_to_fit=existing_alignment.shrink_to_fit,
        indent=existing_alignment.indent
    )

    cell.font = new_font
    cell.alignment = new_alignment

def copy_headers(start_row):
    for i in range(1, 5):
        sheet.row_dimensions[start_row + i - 1].height = sheet.row_dimensions[i].height
        for j in range(1, 6):
            source_cell = sheet.cell(row=i, column=j)
            target_cell = sheet.cell(row=start_row + i - 1, column=j)
            target_cell.value = source_cell.value
            if source_cell.has_style:
                target_cell._style = source_cell._style

def copy_styles(start_row, end_row, template_start=6):
    for i in range(start_row, end_row + 1):
        template_row = template_start + ((i - start_row) % 31)
        sheet.row_dimensions[i].height = sheet.row_dimensions[template_row].height
        for j in range(1, 6):
            source_cell = sheet.cell(row=template_row, column=j)
            target_cell = sheet.cell(row=i, column=j)
            if source_cell.has_style:
                target_cell._style = source_cell._style
            center_and_bold_cell(target_cell)

def merge_cells_with_original_font_size(sheet, start_row, start_col, end_row, end_col):
    original_cell = sheet.cell(row=start_row, column=start_col)
    original_font = original_cell.font
    original_font_size = original_font.size
    
    # Define the border
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Fusion des cellules
    sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
    merged_cell = sheet.cell(row=start_row, column=start_col)

    # Appliquer les styles
    new_font = Font(
        name=original_font.name,
        size=original_font_size,
        bold=True,
        italic=original_font.italic,
        vertAlign=original_font.vertAlign,
        underline=original_font.underline,
        strike=original_font.strike,
        color=original_font.color
    )

    new_alignment = Alignment(
        horizontal='center',
        vertical='center',
        text_rotation=original_cell.alignment.text_rotation,
        wrap_text=original_cell.alignment.wrap_text,
        shrink_to_fit=original_cell.alignment.shrink_to_fit,
        indent=original_cell.alignment.indent
    )

    merged_cell.font = new_font
    merged_cell.alignment = new_alignment
    merged_cell.border = border

    # Réappliquer les styles après fusion pour s'assurer qu'ils sont bien appliqués
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.font = new_font
            cell.alignment = new_alignment
            cell.border = border

# Fill the Excel file
row_index = 1
data_per_page = 25
total_data = len(dsp_data)
pages_needed = (total_data - 1) // data_per_page + 1

ordre = 1  # Initialize ordre outside the loop

for page in range(pages_needed):
    # Copy headers
    copy_headers(row_index)

    # Write specific header data
    sheet.cell(row=row_index, column=1, value=f"لائحة الإعفاءات للمترشحين الأحرار لاجتياز اختبارات التربية البدنية لموسم {annee}")
    sheet.cell(row=row_index + 1, column=5, value=centre)
    
    # Write data
    data_start = row_index + 4
    data_end = min(data_start + data_per_page - 1, data_start + total_data - page * data_per_page - 1)

    current_jour = None
    jour_start = data_start
    current_serie = None
    serie_start = data_start

    for j, row_data in enumerate(dsp_data[page * data_per_page: (page + 1) * data_per_page], start=data_start):
        jour, massar, nom, serie = row_data
        
        if jour != current_jour:
            if current_jour is not None:
                # Merge cells for the previous jour
                merge_cells_with_original_font_size(sheet, jour_start, 1, j-1, 1)
                
                # Merge cells for the last serie of the previous jour
                if current_serie is not None:
                    merge_cells_with_original_font_size(sheet, serie_start, 5, j-1, 5)
            
            current_jour = jour
            jour_start = j
            current_serie = None
            serie_start = j
        
        if serie != current_serie:
            if current_serie is not None:
                # Merge cells for the previous serie
                merge_cells_with_original_font_size(sheet, serie_start, 5, j-1, 5)
            current_serie = serie
            serie_start = j
        
        sheet.cell(row=j, column=1, value=jour)
        sheet.cell(row=j, column=2, value=ordre)
        sheet.cell(row=j, column=3, value=massar)
        sheet.cell(row=j, column=4, value=nom)
        sheet.cell(row=j, column=5, value=serie)
        
        for col in range(1, 6):
            center_and_bold_cell(sheet.cell(row=j, column=col))
        
        ordre += 1

    # Merge cells for the last jour on the page
    merge_cells_with_original_font_size(sheet, jour_start, 1, j, 1)

    # Merge cells for the last serie on the page
    if current_serie is not None:
        merge_cells_with_original_font_size(sheet, serie_start, 5, j, 5)

    # Copy styles for the data rows
    copy_styles(data_start, data_end)

    # Add page break after every page except the last one
    if page < pages_needed - 1:
        sheet.row_breaks.append(Break(data_end))

    # Move to the next section
    row_index = data_end + 1

# Save the filled Excel file
workbook.save(output_excel_path)

# Close the database connection
conn.close()

# Convert the Excel file to PDF
def convert_excel_to_pdf(excel_path, pdf_path):
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    wb = excel.Workbooks.Open(excel_path)
    wb.ExportAsFixedFormat(0, pdf_path)
    wb.Close()
    excel.Quit()

# Open file dialog to save PDF
root = Tk()
root.withdraw()  # Hide the root window
pdf_path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")])

if pdf_path:
    convert_excel_to_pdf(output_excel_path, pdf_path)
    messagebox.showinfo("ممتاز", f"تم تحميل لوائح الإعفاءات بنجاح")
else:
    messagebox.showwarning("خطأ", "حدث خطأ أثناء التحميل")
