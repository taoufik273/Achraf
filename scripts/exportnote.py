import sqlite3
import os
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from tkinter import Tk, filedialog, messagebox

def get_absolute_path(*relative_path):
    script_dir = os.path.dirname(__file__)
    return os.path.join(script_dir, *relative_path)

# Paths
db_path = get_absolute_path('..', 'data', 'saisie.db')
input_excel_path = get_absolute_path('..', 'INPUT', 'export_notes.xlsx')

# Check if database file exists
if not os.path.exists(db_path):
    raise FileNotFoundError(f"Database file not found: {db_path}")

# Connect to the database
conn = sqlite3.connect(db_path)
cursor = conn.cursor()

# Load the Excel template
if not os.path.exists(input_excel_path):
    raise FileNotFoundError(f"Excel template file not found: {input_excel_path}")

workbook = load_workbook(input_excel_path)
sheet = workbook.active

# Fetch data from the index2 table
cursor.execute("SELECT section, centre FROM index2")
index2_data = cursor.fetchone()
section, centre = index2_data[0], index2_data[1]

# Write section and centre to specific cells
sheet["F3"] = section
sheet["G6"] = centre

# Fetch data from the notes table
cursor.execute("SELECT numero, ordre, massar, cin, sexe, nom, serie, note, absent, dispence FROM notes")
notes_data = cursor.fetchall()

# Style to copy from row 9
def copy_styles(source_row, target_row):
    for col in range(1, 11):
        source_cell = sheet.cell(row=source_row, column=col)
        target_cell = sheet.cell(row=target_row, column=col)
        if source_cell.has_style:
            target_cell._style = source_cell._style

# Write data to the Excel file starting from row 9
start_row = 9
for row_index, row_data in enumerate(notes_data, start=start_row):
    for col_index, cell_value in enumerate(row_data, start=1):
        sheet.cell(row=row_index, column=col_index, value=cell_value)
    copy_styles(start_row, row_index)

# Open a file dialog to choose the save location
root = Tk()
root.withdraw()  # Hide the root window
output_excel_path = filedialog.asksaveasfilename(
    defaultextension=".xlsx",
    filetypes=[("Excel files", "*.xlsx")],
    title="Save the Excel file"
)

if output_excel_path:
    # Save the filled Excel file
    workbook.save(output_excel_path)
    messagebox.showinfo("ممتاز", "تم تصدير النقط بنجاح")
else:
    messagebox.showwarning("Cancelled", "Save operation was cancelled.")

# Close the database connection
conn.close()
